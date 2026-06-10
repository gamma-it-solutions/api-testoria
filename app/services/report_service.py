import io
import logging
import math
from collections.abc import Sequence
from datetime import UTC, datetime, timedelta

import openpyxl
from sqlalchemy import and_, case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.core import storage
from app.core.exceptions import NotFoundError
from app.models.mixins import not_deleted
from app.models.project import Project
from app.models.test_case import TestCase
from app.models.test_result import TestResult
from app.models.test_run import TestRun
from app.models.test_suite import TestSuite
from app.schemas.report import (
    CrossProjectReportAnalyticsResponse,
    CustomReportRequest,
    CustomReportResponse,
    CustomReportRow,
    DashboardResponse,
    MetricsDataPoint,
    MetricsResponse,
    PerProjectAnalyticsRow,
    ProjectReportAnalyticsResponse,
    ReportAnalyticsSummary,
    RunAnalyticsItem,
    RunReportAttachment,
    RunReportCaseResult,
    RunReportResponse,
    RunSummary,
    TestCaseDistribution,
    TrendPoint,
)
from app.services import test_run_service
from app.utils import stats

logger = logging.getLogger(__name__)

_RUN_STATUS_KEYS = ("passed", "failed", "blocked", "no_run")


async def _get_project(db: AsyncSession, project_id: int) -> Project:
    result = await db.execute(
        select(Project).where(Project.id == project_id, not_deleted(Project))
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise NotFoundError(f"Project {project_id} not found")
    return project


async def _get_run(db: AsyncSession, run_id: int) -> TestRun:
    result = await db.execute(
        select(TestRun).where(TestRun.id == run_id, not_deleted(TestRun))
    )
    run = result.scalar_one_or_none()
    if run is None:
        raise NotFoundError(f"TestRun {run_id} not found")
    return run


async def _aggregate_run_status_counts(
    db: AsyncSession, run_ids: Sequence[int]
) -> dict[int, dict[str, int]]:
    """Return per-run status counts in a single grouped query.

    Each value is a dict containing `passed`, `failed`, `blocked`, `no_run`,
    and `total` keys (all zero-filled for runs with no results).
    """
    counts: dict[int, dict[str, int]] = {
        run_id: dict.fromkeys((*_RUN_STATUS_KEYS, "total"), 0) for run_id in run_ids
    }
    if not run_ids:
        return counts

    rows = (
        await db.execute(
            select(
                TestResult.test_run_id,
                TestResult.status,
                func.count(TestResult.id),
            )
            .where(
                TestResult.test_run_id.in_(run_ids),
                not_deleted(TestResult),
            )
            .group_by(TestResult.test_run_id, TestResult.status)
        )
    ).all()

    for run_id, status, count in rows:
        bucket = counts[run_id]
        if status in _RUN_STATUS_KEYS:
            bucket[status] = count
        bucket["total"] += count
    return counts


async def get_dashboard(db: AsyncSession, project_id: int) -> DashboardResponse:
    await _get_project(db, project_id)

    # Total suites
    total_suites: int = (
        await db.execute(
            select(func.count(TestSuite.id)).where(
                TestSuite.project_id == project_id, not_deleted(TestSuite)
            )
        )
    ).scalar_one()

    # Total cases
    total_cases: int = (
        await db.execute(
            select(func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id == project_id,
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
        )
    ).scalar_one()

    # Total runs
    total_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id == project_id, not_deleted(TestRun)
            )
        )
    ).scalar_one()

    # Active runs (work in flight: planned or active)
    active_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id == project_id,
                TestRun.status.in_(["planned", "active"]),
                not_deleted(TestRun),
            )
        )
    ).scalar_one()

    # Result distribution — completed runs only. Pass-rate is about finished
    # work, so in-flight runs don't leak into the KPI.
    dist_rows = (
        await db.execute(
            select(TestResult.status, func.count(TestResult.id))
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id == project_id,
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
            .group_by(TestResult.status)
        )
    ).all()
    result_distribution: dict[str, int] = {row[0]: row[1] for row in dist_rows}

    total_results = sum(result_distribution.values())
    passed = result_distribution.get("passed", 0)
    pass_rate = stats.round_ratio(stats.pass_rate(passed, total_results)) or 0.0

    # Recent runs (last 5) with pass/fail counts
    recent_runs_result = await db.execute(
        select(TestRun)
        .where(TestRun.project_id == project_id, not_deleted(TestRun))
        .order_by(TestRun.created_at.desc())
        .limit(5)
    )
    recent_run_objs = list(recent_runs_result.scalars().all())
    recent_counts = await _aggregate_run_status_counts(
        db, [run.id for run in recent_run_objs]
    )

    recent_runs: list[RunSummary] = [
        RunSummary(
            id=run.id,
            name=run.name,
            status=run.status,
            passed=recent_counts[run.id]["passed"],
            failed=recent_counts[run.id]["failed"],
            blocked=recent_counts[run.id]["blocked"],
            no_run=recent_counts[run.id]["no_run"],
            total=recent_counts[run.id]["total"],
            created_at=run.created_at,
        )
        for run in recent_run_objs
    ]

    return DashboardResponse(
        total_test_cases=total_cases,
        total_test_runs=total_runs,
        total_test_suites=total_suites,
        pass_rate=pass_rate,
        active_runs=active_runs,
        recent_runs=recent_runs,
        result_distribution=result_distribution,
    )


def _pass_rate(passed: int, total: int) -> float | None:
    """Rounded pass-rate for response fields (plan 044). Use raw
    `stats.pass_rate` when feeding values into a mean — round once at the end."""
    return stats.round_ratio(stats.pass_rate(passed, total))


async def get_report_analytics(
    db: AsyncSession,
    project_id: int,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    run_status: str | None = None,
    include_trend: bool = True,
) -> ProjectReportAnalyticsResponse:
    """Aggregate everything the Reports & Analytics dashboard needs.

    Filters `runs` and `trend` by the window; `summary` counts are
    project-wide so the totals widget stays stable regardless of the
    window chosen.
    """
    await _get_project(db, project_id)

    # Project-wide counts (not filtered by window)
    total_suites: int = (
        await db.execute(
            select(func.count(TestSuite.id)).where(
                TestSuite.project_id == project_id, not_deleted(TestSuite)
            )
        )
    ).scalar_one()
    total_cases: int = (
        await db.execute(
            select(func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id == project_id,
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
        )
    ).scalar_one()
    total_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id == project_id, not_deleted(TestRun)
            )
        )
    ).scalar_one()
    active_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id == project_id,
                TestRun.status.in_(["planned", "active"]),
                not_deleted(TestRun),
            )
        )
    ).scalar_one()

    # Summary result distribution — completed runs only, across the whole
    # project (not window-filtered). Counts are sums of result rows by status.
    dist_rows = (
        await db.execute(
            select(TestResult.status, func.count(TestResult.id))
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id == project_id,
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
            .group_by(TestResult.status)
        )
    ).all()
    result_distribution: dict[str, int] = {row[0]: row[1] for row in dist_rows}
    total_results = sum(result_distribution.values())

    # Summary overall pass rate (plan 041): mean of each completed run's own
    # pass rate. Delegates to `test_run_service.batch_run_progress` so the
    # denominator matches the run-list / Dashboard definition
    # (`passed / max(cases_in_scope, tested)`) — otherwise a run with untested
    # cases would show 100% here and <100% on the Dashboard.
    completed_runs_for_summary = list(
        (
            await db.execute(
                select(TestRun).where(
                    TestRun.project_id == project_id,
                    TestRun.status == "completed",
                    not_deleted(TestRun),
                )
            )
        )
        .scalars()
        .all()
    )
    progress_by_run = await test_run_service.batch_run_progress(
        db, completed_runs_for_summary
    )
    run_rates = [
        p.pass_rate for p in progress_by_run.values() if p.pass_rate is not None
    ]
    overall_pass_rate = stats.round_ratio(
        sum(run_rates) / len(run_rates) if run_rates else 0.0
    ) or 0.0

    # Runs in window (completed_at, falling back to created_at)
    effective_date = func.coalesce(TestRun.completed_at, TestRun.created_at)
    run_query = select(TestRun).where(
        TestRun.project_id == project_id, not_deleted(TestRun)
    )
    if run_status is not None:
        run_query = run_query.where(TestRun.status == run_status)
    if date_from is not None:
        run_query = run_query.where(effective_date >= date_from)
    if date_to is not None:
        run_query = run_query.where(effective_date <= date_to)
    run_query = run_query.order_by(TestRun.created_at.desc())

    run_objs = list((await db.execute(run_query)).scalars().all())
    run_counts = await _aggregate_run_status_counts(db, [r.id for r in run_objs])

    runs: list[RunAnalyticsItem] = [
        RunAnalyticsItem(
            id=run.id,
            project_id=run.project_id,
            name=run.name,
            status=run.status,
            milestone_id=run.milestone_id,
            assigned_to=run.assigned_to,
            passed=run_counts[run.id]["passed"],
            failed=run_counts[run.id]["failed"],
            blocked=run_counts[run.id]["blocked"],
            no_run=run_counts[run.id]["no_run"],
            total=run_counts[run.id]["total"],
            pass_rate=_pass_rate(
                run_counts[run.id]["passed"], run_counts[run.id]["total"]
            ),
            created_at=run.created_at,
            completed_at=run.completed_at,
        )
        for run in run_objs
    ]

    # Test case distribution (project-wide)
    priority_rows = (
        await db.execute(
            select(TestCase.priority, func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id == project_id,
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
            .group_by(TestCase.priority)
        )
    ).all()
    type_rows = (
        await db.execute(
            select(TestCase.type, func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id == project_id,
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
            .group_by(TestCase.type)
        )
    ).all()
    # Automation coverage is keyed on the user-facing `type` flag ("manual" vs
    # "automated"), not on whether `automation_id` happens to be linked. A
    # case marked automated without an automation_id wire-up still counts as
    # automated for the coverage KPI.
    automated_count: int = (
        await db.execute(
            select(func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id == project_id,
                not_deleted(TestSuite),
                not_deleted(TestCase),
                TestCase.type == "automated",
            )
        )
    ).scalar_one()
    manual_count = total_cases - automated_count

    distribution = TestCaseDistribution(
        by_priority={row[0]: row[1] for row in priority_rows},
        by_type={row[0]: row[1] for row in type_rows},
        by_automation={"automated": automated_count, "manual": manual_count},
    )

    # Trend series (results grouped by day within window)
    trend: list[TrendPoint] = []
    if include_trend:
        trend_query = (
            select(
                func.date(TestResult.tested_at).label("day"),
                func.count(TestResult.id).label("total"),
                func.count(
                    case((TestResult.status == "passed", TestResult.id))
                ).label("passed"),
                func.count(
                    case((TestResult.status == "failed", TestResult.id))
                ).label("failed"),
                func.count(
                    case((TestResult.status == "blocked", TestResult.id))
                ).label("blocked"),
                func.count(
                    case((TestResult.status == "no_run", TestResult.id))
                ).label("no_run"),
            )
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id == project_id,
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
        )
        if date_from is not None:
            trend_query = trend_query.where(TestResult.tested_at >= date_from)
        if date_to is not None:
            trend_query = trend_query.where(TestResult.tested_at <= date_to)
        trend_query = trend_query.group_by(func.date(TestResult.tested_at)).order_by(
            func.date(TestResult.tested_at)
        )

        day_rows = (await db.execute(trend_query)).all()
        by_day = {
            str(row.day): TrendPoint(
                date=str(row.day),
                passed=row.passed,
                failed=row.failed,
                blocked=row.blocked,
                no_run=row.no_run,
                total=row.total,
                pass_rate=_pass_rate(row.passed, row.total),
            )
            for row in day_rows
        }
        trend = _zero_fill_trend(by_day, date_from, date_to)

    summary = ReportAnalyticsSummary(
        total_test_cases=total_cases,
        total_test_suites=total_suites,
        total_test_runs=total_runs,
        active_runs=active_runs,
        overall_pass_rate=overall_pass_rate,
        total_results=total_results,
        result_distribution=result_distribution,
    )

    return ProjectReportAnalyticsResponse(
        project_id=project_id,
        date_from=date_from,
        date_to=date_to,
        summary=summary,
        runs=runs,
        test_case_distribution=distribution,
        trend=trend,
    )


async def _resolve_project_scope(
    db: AsyncSession,
    project_ids: Sequence[int] | None,
    include_archived: bool,
) -> list[Project]:
    """Resolve the project set the cross-project analytics query operates on.

    `project_ids=None` → every visible project (today: every non-deleted,
    optionally non-archived). When the caller supplies an explicit list,
    unknown / soft-deleted ids are silently dropped — matching `/projects/stats`
    behaviour and keeping the endpoint robust to stale ids in the URL.
    """
    query = select(Project).where(not_deleted(Project))
    if not include_archived:
        query = query.where(Project.is_archived.is_(False))
    if project_ids is not None:
        if not project_ids:
            return []
        query = query.where(Project.id.in_(project_ids))
    query = query.order_by(Project.id)
    return list((await db.execute(query)).scalars().all())


async def get_cross_project_report_analytics(
    db: AsyncSession,
    project_ids: Sequence[int] | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    run_status: str | None = None,
    include_trend: bool = True,
    include_archived: bool = False,
) -> CrossProjectReportAnalyticsResponse:
    """Cross-project mirror of `get_report_analytics`.

    Same aggregation rules as the per-project endpoint (plans 035/039/041):
    summary pass rate is the mean of every completed run's own pass rate
    across the resolved project set, distributions sum across projects,
    trend is a single zero-filled daily series.
    """
    projects = await _resolve_project_scope(db, project_ids, include_archived)
    project_id_list = [p.id for p in projects]
    project_name_by_id = {p.id: p.name for p in projects}
    project_archived_by_id = {p.id: p.is_archived for p in projects}

    # Empty scope short-circuit — nothing to count, but echo the empty payload
    # so the UI can render a consistent "no data" state.
    if not project_id_list:
        empty_summary = ReportAnalyticsSummary(
            total_test_cases=0,
            total_test_suites=0,
            total_test_runs=0,
            active_runs=0,
            overall_pass_rate=0.0,
            total_results=0,
            result_distribution={},
        )
        empty_distribution = TestCaseDistribution(
            by_priority={},
            by_type={},
            by_automation={"automated": 0, "manual": 0},
        )
        empty_trend = (
            _zero_fill_trend({}, date_from, date_to) if include_trend else []
        )
        return CrossProjectReportAnalyticsResponse(
            project_ids=list(project_ids) if project_ids is not None else None,
            date_from=date_from,
            date_to=date_to,
            summary=empty_summary,
            runs=[],
            test_case_distribution=empty_distribution,
            trend=empty_trend,
            per_project=[],
        )

    # Scope-wide totals (live rows, all projects in scope, not date-filtered)
    total_suites: int = (
        await db.execute(
            select(func.count(TestSuite.id)).where(
                TestSuite.project_id.in_(project_id_list), not_deleted(TestSuite)
            )
        )
    ).scalar_one()
    total_cases: int = (
        await db.execute(
            select(func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id.in_(project_id_list),
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
        )
    ).scalar_one()
    total_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id.in_(project_id_list), not_deleted(TestRun)
            )
        )
    ).scalar_one()
    active_runs: int = (
        await db.execute(
            select(func.count(TestRun.id)).where(
                TestRun.project_id.in_(project_id_list),
                TestRun.status.in_(["planned", "active"]),
                not_deleted(TestRun),
            )
        )
    ).scalar_one()

    # Result distribution — completed runs only, across the whole scope.
    dist_rows = (
        await db.execute(
            select(TestResult.status, func.count(TestResult.id))
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id.in_(project_id_list),
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
            .group_by(TestResult.status)
        )
    ).all()
    result_distribution: dict[str, int] = {row[0]: row[1] for row in dist_rows}
    total_results = sum(result_distribution.values())

    # Overall pass rate (plan 041): mean of every completed run's own pass rate
    # across the whole scope. Per-project breakdown rows apply the same rule
    # per project so each row matches `/projects/stats` for that project.
    completed_runs = list(
        (
            await db.execute(
                select(TestRun).where(
                    TestRun.project_id.in_(project_id_list),
                    TestRun.status == "completed",
                    not_deleted(TestRun),
                )
            )
        )
        .scalars()
        .all()
    )
    progress_by_run = await test_run_service.batch_run_progress(db, completed_runs)
    run_rates_global = [
        p.pass_rate for p in progress_by_run.values() if p.pass_rate is not None
    ]
    overall_pass_rate = stats.round_ratio(
        sum(run_rates_global) / len(run_rates_global) if run_rates_global else 0.0
    ) or 0.0

    # Per-project rates from the same progress map (no extra DB round-trip).
    rates_by_project: dict[int, list[float]] = {pid: [] for pid in project_id_list}
    for run in completed_runs:
        progress = progress_by_run.get(run.id)
        if progress is None or progress.pass_rate is None:
            continue
        rates_by_project[run.project_id].append(progress.pass_rate)

    # Per-project totals — single grouped query each.
    runs_per_project_rows = (
        await db.execute(
            select(TestRun.project_id, func.count(TestRun.id))
            .where(
                TestRun.project_id.in_(project_id_list), not_deleted(TestRun)
            )
            .group_by(TestRun.project_id)
        )
    ).all()
    runs_per_project: dict[int, int] = {pid: 0 for pid in project_id_list}
    for pid, count in runs_per_project_rows:
        runs_per_project[pid] = count

    completed_runs_per_project: dict[int, int] = {pid: 0 for pid in project_id_list}
    for run in completed_runs:
        completed_runs_per_project[run.project_id] += 1

    results_per_project_rows = (
        await db.execute(
            select(TestRun.project_id, func.count(TestResult.id))
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id.in_(project_id_list),
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
            .group_by(TestRun.project_id)
        )
    ).all()
    results_per_project: dict[int, int] = {pid: 0 for pid in project_id_list}
    for pid, count in results_per_project_rows:
        results_per_project[pid] = count

    per_project = [
        PerProjectAnalyticsRow(
            project_id=pid,
            project_name=project_name_by_id[pid],
            is_archived=project_archived_by_id[pid],
            total_test_runs=runs_per_project[pid],
            completed_runs=completed_runs_per_project[pid],
            overall_pass_rate=stats.round_ratio(
                sum(rates_by_project[pid]) / len(rates_by_project[pid])
                if rates_by_project[pid]
                else None
            ),
            total_results=results_per_project[pid],
        )
        for pid in project_id_list
    ]

    # Runs in window (filtered by completed_at, falling back to created_at).
    effective_date = func.coalesce(TestRun.completed_at, TestRun.created_at)
    run_query = select(TestRun).where(
        TestRun.project_id.in_(project_id_list), not_deleted(TestRun)
    )
    if run_status is not None:
        run_query = run_query.where(TestRun.status == run_status)
    if date_from is not None:
        run_query = run_query.where(effective_date >= date_from)
    if date_to is not None:
        run_query = run_query.where(effective_date <= date_to)
    run_query = run_query.order_by(TestRun.created_at.desc())

    run_objs = list((await db.execute(run_query)).scalars().all())
    run_counts = await _aggregate_run_status_counts(db, [r.id for r in run_objs])

    runs: list[RunAnalyticsItem] = [
        RunAnalyticsItem(
            id=run.id,
            project_id=run.project_id,
            project_name=project_name_by_id.get(run.project_id),
            name=run.name,
            status=run.status,
            milestone_id=run.milestone_id,
            assigned_to=run.assigned_to,
            passed=run_counts[run.id]["passed"],
            failed=run_counts[run.id]["failed"],
            blocked=run_counts[run.id]["blocked"],
            no_run=run_counts[run.id]["no_run"],
            total=run_counts[run.id]["total"],
            pass_rate=_pass_rate(
                run_counts[run.id]["passed"], run_counts[run.id]["total"]
            ),
            created_at=run.created_at,
            completed_at=run.completed_at,
        )
        for run in run_objs
    ]

    # Test case distribution (scope-wide).
    priority_rows = (
        await db.execute(
            select(TestCase.priority, func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id.in_(project_id_list),
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
            .group_by(TestCase.priority)
        )
    ).all()
    type_rows = (
        await db.execute(
            select(TestCase.type, func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id.in_(project_id_list),
                not_deleted(TestSuite),
                not_deleted(TestCase),
            )
            .group_by(TestCase.type)
        )
    ).all()
    automated_count: int = (
        await db.execute(
            select(func.count(TestCase.id))
            .join(TestSuite, TestCase.suite_id == TestSuite.id)
            .where(
                TestSuite.project_id.in_(project_id_list),
                not_deleted(TestSuite),
                not_deleted(TestCase),
                TestCase.type == "automated",
            )
        )
    ).scalar_one()
    manual_count = total_cases - automated_count

    distribution = TestCaseDistribution(
        by_priority={row[0]: row[1] for row in priority_rows},
        by_type={row[0]: row[1] for row in type_rows},
        by_automation={"automated": automated_count, "manual": manual_count},
    )

    # Trend series — one aggregated daily series across all in-scope projects.
    trend: list[TrendPoint] = []
    if include_trend:
        trend_query = (
            select(
                func.date(TestResult.tested_at).label("day"),
                func.count(TestResult.id).label("total"),
                func.count(
                    case((TestResult.status == "passed", TestResult.id))
                ).label("passed"),
                func.count(
                    case((TestResult.status == "failed", TestResult.id))
                ).label("failed"),
                func.count(
                    case((TestResult.status == "blocked", TestResult.id))
                ).label("blocked"),
                func.count(
                    case((TestResult.status == "no_run", TestResult.id))
                ).label("no_run"),
            )
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id.in_(project_id_list),
                TestRun.status == "completed",
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
        )
        if date_from is not None:
            trend_query = trend_query.where(TestResult.tested_at >= date_from)
        if date_to is not None:
            trend_query = trend_query.where(TestResult.tested_at <= date_to)
        trend_query = trend_query.group_by(func.date(TestResult.tested_at)).order_by(
            func.date(TestResult.tested_at)
        )

        day_rows = (await db.execute(trend_query)).all()
        by_day = {
            str(row.day): TrendPoint(
                date=str(row.day),
                passed=row.passed,
                failed=row.failed,
                blocked=row.blocked,
                no_run=row.no_run,
                total=row.total,
                pass_rate=_pass_rate(row.passed, row.total),
            )
            for row in day_rows
        }
        trend = _zero_fill_trend(by_day, date_from, date_to)

    summary = ReportAnalyticsSummary(
        total_test_cases=total_cases,
        total_test_suites=total_suites,
        total_test_runs=total_runs,
        active_runs=active_runs,
        overall_pass_rate=overall_pass_rate,
        total_results=total_results,
        result_distribution=result_distribution,
    )

    return CrossProjectReportAnalyticsResponse(
        project_ids=list(project_ids) if project_ids is not None else None,
        date_from=date_from,
        date_to=date_to,
        summary=summary,
        runs=runs,
        test_case_distribution=distribution,
        trend=trend,
        per_project=per_project,
    )


def _zero_fill_trend(
    by_day: dict[str, TrendPoint],
    date_from: datetime | None,
    date_to: datetime | None,
) -> list[TrendPoint]:
    """Emit a continuous day-by-day series when a window is specified."""
    if date_from is None or date_to is None:
        return [by_day[k] for k in sorted(by_day.keys())]

    start = date_from.date()
    end = date_to.date()
    if start > end:
        return []

    out: list[TrendPoint] = []
    day = start
    while day <= end:
        key = day.isoformat()
        if key in by_day:
            out.append(by_day[key])
        else:
            out.append(
                TrendPoint(
                    date=key,
                    passed=0,
                    failed=0,
                    blocked=0,
                    no_run=0,
                    total=0,
                    pass_rate=None,
                )
            )
        day = day + timedelta(days=1)
    return out


async def get_run_report(db: AsyncSession, run_id: int) -> RunReportResponse:
    run = await _get_run(db, run_id)

    # Fetch all cases in scope with their results for this run
    cases_query = (
        select(TestCase, TestResult, TestSuite.name.label("suite_name"))
        .join(TestSuite, TestCase.suite_id == TestSuite.id)
        .outerjoin(
            TestResult,
            and_(
                TestResult.test_case_id == TestCase.id,
                TestResult.test_run_id == run_id,
                not_deleted(TestResult),
            ),
        )
        .options(selectinload(TestResult.attachments))
        .where(
            TestSuite.project_id == run.project_id,
            not_deleted(TestSuite),
            not_deleted(TestCase),
        )
    )
    if run.suite_id is not None:
        cases_query = cases_query.where(TestCase.suite_id == run.suite_id)
    cases_query = cases_query.order_by(TestCase.id)

    rows = (await db.execute(cases_query)).all()

    cases: list[RunReportCaseResult] = []
    passed = failed = blocked = no_run = 0
    for tc, res, suite_name in rows:
        status = res.status if res else None
        attachments_for_report: list[RunReportAttachment] = []
        if res is not None and res.attachments:
            limit = settings.REPORT_ATTACHMENT_MAX_PER_RESULT
            for att in res.attachments[:limit]:
                url = ""
                if att.storage_backend == "s3":
                    try:
                        url = storage.generate_presigned_url_sync(
                            att.object_key,
                            ttl_seconds=settings.S3_REPORT_PRESIGN_TTL_SECONDS,
                        )
                    except Exception as err:  # noqa: BLE001
                        logger.warning(
                            "Failed to presign %s for report: %s",
                            att.object_key,
                            err,
                        )
                attachments_for_report.append(
                    RunReportAttachment(
                        id=att.id,
                        filename=att.filename,
                        mime_type=att.mime_type,
                        url=url,
                        object_key=att.object_key,
                        storage_backend=att.storage_backend,
                    )
                )
        cases.append(
            RunReportCaseResult(
                case_id=tc.id,
                case_title=tc.title,
                suite_name=suite_name,
                priority=tc.priority,
                status=status,
                comment=res.comment if res else None,
                tested_by=res.tested_by if res else None,
                tested_at=res.tested_at if res else None,
                execution_time=res.execution_time if res else None,
                attachments=attachments_for_report,
            )
        )
        if status == "passed":
            passed += 1
        elif status == "failed":
            failed += 1
        elif status == "blocked":
            blocked += 1
        else:
            # Both explicit `no_run` and missing result → no_run bucket.
            no_run += 1

    total = len(cases)
    pass_rate = _pass_rate(passed, total)

    return RunReportResponse(
        run_id=run.id,
        run_name=run.name,
        run_status=run.status,
        project_id=run.project_id,
        created_at=run.created_at,
        completed_at=run.completed_at,
        passed=passed,
        failed=failed,
        blocked=blocked,
        no_run=no_run,
        total=total,
        pass_rate=pass_rate,
        cases=cases,
    )


async def get_project_metrics(
    db: AsyncSession, project_id: int, days: int = 30
) -> MetricsResponse:
    await _get_project(db, project_id)

    cutoff = datetime.now(UTC) - timedelta(days=days)

    # Group results by date, counting statuses
    rows = (
        await db.execute(
            select(
                func.date(TestResult.tested_at).label("day"),
                func.count(TestResult.id).label("total"),
                func.count(case((TestResult.status == "passed", TestResult.id))).label(
                    "passed"
                ),
                func.count(case((TestResult.status == "failed", TestResult.id))).label(
                    "failed"
                ),
                func.count(case((TestResult.status == "blocked", TestResult.id))).label(
                    "blocked"
                ),
                func.count(case((TestResult.status == "no_run", TestResult.id))).label(
                    "no_run"
                ),
            )
            .join(TestRun, TestResult.test_run_id == TestRun.id)
            .where(
                TestRun.project_id == project_id,
                TestResult.tested_at >= cutoff,
                not_deleted(TestRun),
                not_deleted(TestResult),
            )
            .group_by(func.date(TestResult.tested_at))
            .order_by(func.date(TestResult.tested_at))
        )
    ).all()

    data: list[MetricsDataPoint] = []
    for row in rows:
        total = row.total
        p = row.passed
        data.append(
            MetricsDataPoint(
                date=str(row.day),
                passed=p,
                failed=row.failed,
                blocked=row.blocked,
                no_run=row.no_run,
                total=total,
                pass_rate=stats.round_ratio(stats.pass_rate(p, total)),
            )
        )

    return MetricsResponse(project_id=project_id, days=days, data=data)


async def run_custom_report(
    db: AsyncSession, filters: CustomReportRequest
) -> CustomReportResponse:
    await _get_project(db, filters.project_id)

    query = (
        select(
            TestResult.id.label("result_id"),
            TestRun.id.label("run_id"),
            TestRun.name.label("run_name"),
            TestCase.id.label("case_id"),
            TestCase.title.label("case_title"),
            TestResult.status,
            TestResult.comment,
            TestResult.tested_by,
            TestResult.tested_at,
        )
        .join(TestRun, TestResult.test_run_id == TestRun.id)
        .join(TestCase, TestResult.test_case_id == TestCase.id)
        .where(
            TestRun.project_id == filters.project_id,
            not_deleted(TestRun),
            not_deleted(TestResult),
            not_deleted(TestCase),
        )
    )

    if filters.suite_id is not None:
        query = query.join(TestSuite, TestCase.suite_id == TestSuite.id).where(
            TestSuite.id == filters.suite_id
        )
    if filters.run_id is not None:
        query = query.where(TestRun.id == filters.run_id)
    if filters.status:
        query = query.where(TestResult.status.in_(filters.status))
    if filters.date_from is not None:
        query = query.where(TestResult.tested_at >= filters.date_from)
    if filters.date_to is not None:
        query = query.where(TestResult.tested_at <= filters.date_to)

    # Count total
    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total: int = count_result.scalar_one()

    # Paginate
    offset = (filters.page - 1) * filters.page_size
    rows = (
        await db.execute(
            query.order_by(TestResult.tested_at.desc())
            .offset(offset)
            .limit(filters.page_size)
        )
    ).all()

    items = [
        CustomReportRow(
            result_id=row.result_id,
            run_id=row.run_id,
            run_name=row.run_name,
            case_id=row.case_id,
            case_title=row.case_title,
            status=row.status,
            comment=row.comment,
            tested_by=row.tested_by,
            tested_at=row.tested_at,
        )
        for row in rows
    ]

    pages = math.ceil(total / filters.page_size) if filters.page_size else 1
    return CustomReportResponse(
        items=items,
        total=total,
        page=filters.page,
        page_size=filters.page_size,
        pages=pages,
    )


async def _fetch_attachment_bytes(
    att: RunReportAttachment,
) -> bytes | None:
    """Download image bytes for report embed. Returns None on failure."""
    if att.storage_backend != "s3" or not att.object_key:
        return None
    try:
        return await storage.get_object(att.object_key)
    except Exception as err:  # noqa: BLE001
        logger.warning(
            "Report embed: failed to fetch %s (%s)", att.object_key, err
        )
        return None


def _downsize_for_embed(image_bytes: bytes, max_width: int = 800) -> bytes:
    """Resize an image to `max_width` pixels wide; return re-encoded bytes.

    Keeps embedded images small so report PDFs / workbooks stay reasonable.
    """
    try:
        from PIL import Image
    except ImportError:
        return image_bytes
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            if img.width <= max_width:
                return image_bytes
            ratio = max_width / img.width
            new_size = (max_width, int(img.height * ratio))
            resized = img.resize(new_size)
            out = io.BytesIO()
            fmt = (img.format or "PNG").upper()
            if fmt == "JPEG" and resized.mode != "RGB":
                resized = resized.convert("RGB")
            resized.save(out, format=fmt if fmt in {"PNG", "JPEG"} else "PNG")
            return out.getvalue()
    except Exception as err:  # noqa: BLE001
        logger.warning("Image downsize failed: %s — using original", err)
        return image_bytes


async def generate_run_report_pdf(db: AsyncSession, run_id: int) -> bytes:
    report = await get_run_report(db, run_id)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image as RLImage,
    )
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements: list[object] = []

    # Title
    title_text = f"Test Run Report: {report.run_name}"
    elements.append(Paragraph(title_text, styles["Title"]))
    elements.append(Spacer(1, 6 * mm))

    # Summary
    summary_lines = [
        f"Status: {report.run_status}",
        f"Total: {report.total}  |  Passed: {report.passed}  |  Failed: {report.failed}"
        f"  |  Blocked: {report.blocked}  |  No-run: {report.no_run}",
        f"Pass rate: {report.pass_rate:.1%}"
        if report.pass_rate is not None
        else "Pass rate: N/A",
    ]
    for line in summary_lines:
        elements.append(Paragraph(line, styles["Normal"]))
    elements.append(Spacer(1, 6 * mm))

    # Results table
    table_data = [["#", "Case", "Priority", "Status", "Comment"]]
    for i, c in enumerate(report.cases, 1):
        table_data.append(
            [
                str(i),
                c.case_title[:60],
                c.priority,
                c.status or "no_run",
                (c.comment or "")[:80],
            ]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F2F2F2")],
                ),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)

    # Per-result screenshot section (attachments embedded inline)
    cases_with_attachments = [c for c in report.cases if c.attachments]
    if cases_with_attachments:
        elements.append(Spacer(1, 8 * mm))
        elements.append(Paragraph("Screenshots", styles["Heading2"]))
        elements.append(Spacer(1, 3 * mm))
        for c in cases_with_attachments:
            elements.append(
                Paragraph(
                    f"<b>{c.case_title[:80]}</b> ({c.status or 'no_run'})",
                    styles["Normal"],
                )
            )
            for att in c.attachments:
                content = await _fetch_attachment_bytes(att)
                if content is None:
                    elements.append(
                        Paragraph(
                            f"[{att.filename} — unavailable]", styles["Normal"]
                        )
                    )
                    continue
                try:
                    bytes_ = _downsize_for_embed(content, max_width=800)
                    img = RLImage(
                        io.BytesIO(bytes_), width=120 * mm, height=80 * mm,
                        kind="proportional",
                    )
                    elements.append(img)
                    elements.append(Spacer(1, 2 * mm))
                except Exception as err:  # noqa: BLE001
                    logger.warning(
                        "Failed to embed %s in PDF: %s", att.filename, err
                    )
                    elements.append(
                        Paragraph(
                            f"[{att.filename} — embed failed]", styles["Normal"]
                        )
                    )
            elements.append(Spacer(1, 4 * mm))

    doc.build(elements)
    return buf.getvalue()


async def generate_run_report_excel(db: AsyncSession, run_id: int) -> bytes:
    from openpyxl.drawing.image import Image as XLImage

    report = await get_run_report(db, run_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Test Results"

    # Header
    ws.append(
        [
            "Case ID",
            "Title",
            "Suite",
            "Priority",
            "Status",
            "Comment",
            "Tested By",
            "Tested At",
            "Execution Time (s)",
            "Screenshots",
        ]
    )

    for c in report.cases:
        screenshot_text = ""
        if c.attachments:
            kept = len(c.attachments)
            screenshot_text = (
                f"{kept} screenshot(s) — see below for inline thumbnails"
            )
        ws.append(
            [
                c.case_id,
                c.case_title,
                c.suite_name or "",
                c.priority,
                c.status or "no_run",
                c.comment or "",
                c.tested_by,
                str(c.tested_at) if c.tested_at else "",
                c.execution_time,
                screenshot_text,
            ]
        )

    # Embed screenshots in a dedicated "Screenshots" sheet so the data table
    # stays tabular. Each case with attachments gets its own block.
    cases_with_attachments = [c for c in report.cases if c.attachments]
    if cases_with_attachments:
        img_ws = wb.create_sheet("Screenshots")
        img_ws.column_dimensions["A"].width = 30
        img_ws.column_dimensions["B"].width = 80
        row_cursor = 1
        for c in cases_with_attachments:
            img_ws.cell(row=row_cursor, column=1, value="Case").font = (
                openpyxl.styles.Font(bold=True)
            )
            img_ws.cell(
                row=row_cursor,
                column=2,
                value=f"{c.case_title} (id={c.case_id})",
            )
            row_cursor += 1
            for att in c.attachments:
                content = await _fetch_attachment_bytes(att)
                if content is None:
                    img_ws.cell(
                        row=row_cursor,
                        column=2,
                        value=f"[{att.filename} — unavailable]",
                    )
                    row_cursor += 1
                    continue
                try:
                    resized = _downsize_for_embed(content, max_width=600)
                    img = XLImage(io.BytesIO(resized))
                    img.width = 300
                    img.height = 200
                    img_ws.row_dimensions[row_cursor].height = 155
                    img_ws.add_image(img, f"B{row_cursor}")
                    img_ws.cell(
                        row=row_cursor, column=1, value=att.filename
                    )
                    row_cursor += 11  # leave vertical room for the image
                except Exception as err:  # noqa: BLE001
                    logger.warning(
                        "Failed to embed %s in Excel: %s", att.filename, err
                    )
                    img_ws.cell(
                        row=row_cursor,
                        column=2,
                        value=f"[{att.filename} — embed failed]",
                    )
                    row_cursor += 1
            row_cursor += 2  # gap between cases

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
